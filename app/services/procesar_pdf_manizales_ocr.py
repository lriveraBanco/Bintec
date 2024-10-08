import os
from openpyxl import load_workbook
from app.utils.pdf_processing_imagen import (
    letra_a_indice_columna,
    preprocesar_imagen,
    procesar_pdf_manizales_ocr,
    limpiar_texto,
    buscar_articulo_44,
    buscar_articulo_92
)

def procesar_pdf_y_guardar_en_excel_manizales(ruta_pdf, municipio):
    """Procesa el PDF y guarda la información relevante en un archivo Excel."""
    ruta_excel = os.path.join('data', 'municipios', municipio, 'plantilla', 'Ejemplos industria y comercio3.xlsx')
    texto_pdf = procesar_pdf_manizales_ocr(ruta_pdf)
    valor_uyt_uvt_articulo_44 = buscar_articulo_44(texto_pdf)
    porcentaje_articulo_92 = buscar_articulo_92(texto_pdf)
    if porcentaje_articulo_92:
        libro = load_workbook(ruta_excel)
        hoja_parametros = libro['Parametros ICA']
        for fila in range(4, hoja_parametros.max_row + 1):
            ciudad = hoja_parametros.cell(row=fila, column=letra_a_indice_columna('D')).value
            if ciudad == "MANIZALES":
                compania = hoja_parametros.cell(row=fila, column=letra_a_indice_columna('A')).value
                if compania in ["Banca", "CFNS", "Valores", "Fiduciaria"]:
                    hoja_parametros.cell(row=fila, column=letra_a_indice_columna('J'), value="0%")
                    print(f"Compañía especial '{compania}' en fila {fila}: Asignado 0%")
                else:
                    hoja_parametros.cell(row=fila, column=letra_a_indice_columna('J'), value=porcentaje_articulo_92)
                    print(f"Porcentaje del Artículo 92 para fila {fila}: {porcentaje_articulo_92}")
        if valor_uyt_uvt_articulo_44:
            for fila in range(4, hoja_parametros.max_row + 1):
                ciudad = hoja_parametros.cell(row=fila, column=letra_a_indice_columna('D')).value
                valor_columna_o = hoja_parametros.cell(row=fila, column=letra_a_indice_columna('O')).value
                if ciudad == "MANIZALES" and valor_columna_o == "Si":
                    hoja_parametros.cell(row=fila, column=letra_a_indice_columna('Q'), value=valor_uyt_uvt_articulo_44)
                    print(f"Valor del Artículo 44 ({valor_uyt_uvt_articulo_44}) colocado en fila {fila} en Manizales.")
        else:
            print("No se encontró un valor válido para el Artículo 44.")           
        libro.save(ruta_excel)
        return {"mensaje": "Datos guardados en Excel", "archivo": ruta_excel}
    else:
        print("No se encontró el ARTÍCULO 92 en el PDF.")