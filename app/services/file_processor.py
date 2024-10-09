import os
from openpyxl import load_workbook
from app.utils.pdf_processing import (
    extraer_texto_pdf,
    buscar_articulo_71,
    buscar_tarifa_2024,
    buscar_articulo_97,
    buscar_articulo_105,
    letra_a_indice_columna,
    buscar_valor_articulo_56
)


def procesar_pdf_y_guardar_en_excel_medellin(ruta_pdf, municipio):
    """Procesa un PDF y guarda la información en un archivo de Excel específico."""
    ruta_excel = os.path.join('data', 'municipios', municipio, 'plantilla', 'Ejemplos industria y comercio3.xlsx')

    texto_pdf = extraer_texto_pdf(ruta_pdf)
    texto_despues_articulo_71 = buscar_articulo_71(texto_pdf)or None
    porcentaje_articulo_97 = buscar_articulo_97(texto_pdf)or None
    porcentaje_articulo_105 = buscar_articulo_105(texto_pdf)or None
    valor_uyt_uvt_articulo_56 = buscar_valor_articulo_56(texto_pdf)or None
    if texto_despues_articulo_71:
        libro = load_workbook(ruta_excel)
        hoja_parametros = libro['Parametros ICA']
        
        for fila in range(4, hoja_parametros.max_row + 1):
            ciudad = hoja_parametros.cell(row=fila, column=letra_a_indice_columna('D')).value
            if ciudad == "MEDELLIN":
                codigo_ciiu = hoja_parametros.cell(row=fila, column=letra_a_indice_columna('AV')).value
                if codigo_ciiu:
                    tarifa_2024 = buscar_tarifa_2024(texto_despues_articulo_71, codigo_ciiu)
                    if tarifa_2024:
                        hoja_parametros.cell(row=fila, column=letra_a_indice_columna('I'), value=tarifa_2024)
                        hoja_parametros.cell(row=fila, column=letra_a_indice_columna('AD'), value=tarifa_2024)
                        print(f"Tarifa del año 2024 para {codigo_ciiu}: {tarifa_2024}")
                    else:
                        print(f"No se encontró tarifa para el código {codigo_ciiu}")

        if porcentaje_articulo_97:
            for fila in range(4, hoja_parametros.max_row + 1):
                ciudad = hoja_parametros.cell(row=fila, column=letra_a_indice_columna('D')).value
                if ciudad == "MEDELLIN":
                    compania = hoja_parametros.cell(row=fila, column=letra_a_indice_columna('A')).value
                    if compania in ["Banca", "CFNS", "Valores", "Fiduciaria"]:
                        hoja_parametros.cell(row=fila, column=letra_a_indice_columna('J'), value="0%")
                        print(f"Compañía especial '{compania}' en fila {fila}: Asignado 0%")
                    else:
                        hoja_parametros.cell(row=fila, column=letra_a_indice_columna('J'), value=porcentaje_articulo_97)
                        print(f"Porcentaje del Artículo 97 para fila {fila}: {porcentaje_articulo_97}")

        if porcentaje_articulo_105:
            for fila in range(4, hoja_parametros.max_row + 1):
                ciudad = hoja_parametros.cell(row=fila, column=letra_a_indice_columna('D')).value
                if ciudad == "MEDELLIN":
                    hoja_parametros.cell(row=fila, column=letra_a_indice_columna('L'), value=porcentaje_articulo_105)
                    print(f"Porcentaje del Artículo 105 para fila {fila}: {porcentaje_articulo_105}")
        if valor_uyt_uvt_articulo_56:
            for fila in range(4, hoja_parametros.max_row + 1):
                ciudad = hoja_parametros.cell(row=fila, column=letra_a_indice_columna('D')).value
                valor_columna_o = hoja_parametros.cell(row=fila, column=letra_a_indice_columna('O')).value
                if ciudad == "MEDELLIN" and valor_columna_o == "Si":
                    hoja_parametros.cell(row=fila, column=letra_a_indice_columna('Q'), value=valor_uyt_uvt_articulo_56)
                    print(f"Valor del Artículo 56 (equivalente a {valor_uyt_uvt_articulo_56}) colocado en fila {fila} en Medellín.")
        else:
            print("No se encontró un valor válido para el Artículo 56.")

        libro.save(ruta_excel)
        return {"mensaje": "Datos guardados en Excel", "archivo": ruta_excel}
    else:
        return {"error": "No se encontró ningun ARTÍCULO en el PDF."}

def letra_a_indice_columna(letra_columna):
    """Convierte la letra de una columna de Excel en un índice numérico."""
    indice = 0
    for i, letra in enumerate(reversed(letra_columna)):
        indice += (ord(letra.upper()) - ord('A') + 1) * (26 ** i)
    return indice
